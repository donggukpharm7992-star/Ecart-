from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "src" / "data" / "inventory.generated.json"
NARCOTIC_INVENTORY_OUT = ROOT / "src" / "data" / "narcoticInventory.generated.json"
HOSPITAL_DRUG_LIST_PATH = ROOT / "\uc57d\uc81c\ud300 \ub77c\ubca8" / "\uc6d0\ub0b4\ubcf4\uc720\uc758\uc57d\ud488\ub9ac\uc2a4\ud2b8.xlsx"
NARCOTIC_STATUS_KEYWORD = "\ube44\uce58\ud5a5\uc815,\ub9c8\uc57d\ud604\ud669"
NARCOTIC_CHECK_SHEET_NAME = "\uc810\uac80"
CONTROLLED_DRUG_PREFIX_RE = re.compile(r"^\[(?:\ub9c8\uc57d|\ud5a5\uc815)\]\s*")


SHEET_ALIAS = {
    "42W": "42",
    "61W": "61",
    "62W": "62",
    "71W": "71",
    "72W": "72",
    "81W": "81",
    "82W": "82",
    "91W": "91",
    "92W": "92",
    "101W": "101",
    "102W": "102",
    "111W": "111",
    "112W": "112",
    "121W": "121",
    "신속대응팀": "RRT",
    "HBEF심혈관조영실": "HBEF",
    "체외순환실": "HEART",
    "외래주사실": "INJ",
    "영상의학과": "DRO",
    "안과": "OT",
    "재활의학과": "RH",
    "산부인과": "OG",
    "난임클리닉": "난임",
    "이비인후과": "OL",
    "피부과": "DM",
    "소화기병검사실": "GICLA",
    "정형외과": "OS",
    "신경과": "NR",
    "비뇨기과": "UR",
}

STOCK_CODE_OVERRIDES = {
    "0.9% NaKCl 20mEq/100ml btl": "XNAK20",
}

CHECKLIST_LABEL_ROWS = {"양호불량"}

STOCK_FIELD_OVERRIDES = {
    "XBPCA5W": {"warning": ""},
    "XEPIN": {"storageType": "ROOM"},
    "XKPHMB": {"warning": "고위험의약품"},
    "XMEXO": {"warning": "유사모양"},
    "XMVH": {"storageType": "REFRIGERATED"},
    "XNA40": {"warning": "고위험의약품"},
}

NARCOTIC_COLD_STORAGE_CODES = {"XLZPAM2", "XLZPAM4", "XKETA5"}
NARCOTIC_ROOM_FLOORS = {
    "INJ": "1\uce35",
    "DREMM": "1\uce35",
    "ER": "1\uce35",
    "GICLA": "2\uce35",
    "HBEF": "2\uce35",
    "MICU": "2\uce35",
    "DSR": "2\uce35",
    "AN": "3\uce35",
    "OR": "3\uce35",
    "SICU": "3\uce35",
    "ADR": "4\uce35",
    "HPC": "4\uce35",
    "\ub09c\uc784": "4\uce35",
    "DRL": "4\uce35",
    "NICU": "4\uce35",
    "42": "4\uce35",
}

ECART_FIELD_OVERRIDES = {
    "XADENO6": {"name": "Adenocor( Adenosin )", "dosage": "6mg/Vial", "quantity": 3},
    "XLID2W": {"name": "2% Lidocaine 400mg", "dosage": "2% 20mL/Vial", "quantity": 2},
    "XNB84": {"name": "Sodium Bicabonate", "dosage": "20mEq/20mL/Amp", "quantity": 10},
    "XNITR10F": {"name": "Nitrolingual 0.1%", "dosage": "10mg/10ml", "quantity": 5},
    "XNS20": {"name": "N/S 20cc", "dosage": "20mL/Amp", "quantity": 3},
    "XCPENIR": {"name": "Peniramin", "dosage": "4mg/2ml/Amp", "quantity": 3},
    "NITR": {"name": "Nitroglycerin(SL)", "dosage": "0.6mg/Tab", "quantity": 3},
}

ECART_NICU_CODE_OVERRIDES = [
    ("Ephedrine", "40mg/Amp", "XEPHE"),
    ("Tropin(Dopamine hcl)", "200mg/Amp", "XDOPA4"),
    ("Digoxin", "\uc8fc:0.25mg/Amp", "XDGX"),
    ("Doburan", "\uc8fc:250mg/Pfs", "XDOB250"),
    ("Magnesium sulfate", "\uc8fc:10% 20ml Amp", "XMGSF10"),
    ("Predisol 125mg inj", "\uc8fc:125mg/2ml/ Vial", "XMPD1W"),
    ("Protamin Sulfate", "\uc8fc:50mg/Amp", "XPROT"),
    ("Isoptin", "\uc8fc:5mg/Amp", "XVERAW"),
    ("Sodium Chloride 3%", "\uc8fc: 3% 500ml/Btl", "X3NACL5"),
    ("Water for Injection", "\uc8fc:1000ml/Bag", "XAQD"),
    ("N/S", "500ml - Bag", "XNS500"),
    ("5% D/W", "500ml - Bag", "XD5W5"),
    ("5% DW 100ml bag", "500ml - Bag", "XD5W5"),
    ("20% D/W", "4g/20ml/Amp", "XD20W20"),
    ("Normal saline", "50ml - Bag", "XNS50C"),
    ("5%DNK 3", "\uc8fc:(Na34,K20)500ml/Bag", "XDNK35"),
]


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    return "" if text in {"#N/A", "None"} else re.sub(r"\s+", " ", text)


def qty(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    text = clean(value)
    if not text:
        return 0
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return int(float(match.group(0))) if match else 0


def allocation_qty(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    text = clean(value)
    if not text:
        return 0
    return int(float(text)) if re.fullmatch(r"-?\d+(?:\.\d+)?", text) else 0


def room_update_date(value: Any) -> str:
    text = clean(value).replace(" ", "")
    match = re.fullmatch(r"(\d{2,4})\.(\d{1,2})\.?(\d{1,2})", text)
    if not match:
        return ""
    year, month, day = match.groups()
    return f"{year[-2:]}.{int(month):02d}.{int(day):02d}"


def collect_room_update_dates(wb: Any) -> dict[str, str]:
    dates: dict[str, str] = {}
    for ws in wb.worksheets[1:]:
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            for value in row:
                date = room_update_date(value)
                if date:
                    dates[ws.title] = date
                    break
            if ws.title in dates:
                break
    return dates


def storage_type(storage: str) -> str:
    text = storage.replace(" ", "")
    if "냉장보관하지" in text:
        return "ROOM"
    if "25℃" in text or "25도" in text:
        return "ROOM"
    if any(token in text for token in ["냉장", "2~8", "2-8", "2∼8", "5±3", "5℃이하", "10℃이하"]):
        return "REFRIGERATED"
    if "차광" in text:
        return "LIGHT_PROTECTED"
    return "ROOM"


def stock_code(raw_code: str, product_name: str) -> str:
    return STOCK_CODE_OVERRIDES.get(product_name, raw_code)


def is_checklist_label_row(text: str) -> bool:
    return re.sub(r"\s+", "", text) in CHECKLIST_LABEL_ROWS


def is_retired_checklist_row(text: str) -> bool:
    return "E-cart" in text and "\uc8fc 2\ud68c" in text and "\uad00\ub9ac\ub300\uc7a5" in text


def find_workbook(keyword: str) -> Path:
    matches = [p for p in ROOT.rglob("*.xlsx") if keyword in p.name and not p.name.startswith("~$")]
    if not matches:
        raise FileNotFoundError(f"Cannot find workbook containing {keyword!r}")
    return matches[0]


def strip_controlled_drug_prefix(name: str) -> str:
    return CONTROLLED_DRUG_PREFIX_RE.sub("", clean(name))


def load_hospital_common_names(hospital_path: Path) -> dict[str, str]:
    wb = openpyxl.load_workbook(hospital_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    headers = [clean(value) for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    header_index = {header: index for index, header in enumerate(headers)}
    code_index = header_index["약품코드"]
    common_name_index = header_index["상용약품명"]
    common_names: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = clean(row[code_index] if len(row) > code_index else "")
        common_name = strip_controlled_drug_prefix(row[common_name_index] if len(row) > common_name_index else "")
        if code and common_name and code not in common_names:
            common_names[code] = common_name
    return common_names


def narcotic_source_updated_at(title: str) -> str:
    match = re.search(r"(\d{4})\ub144\s*(\d{1,2})\uc6d4\s*(\d{1,2})\uc77c", title)
    if not match:
        return ""
    year, month, day = match.groups()
    return f"{year[-2:]}.{int(month):02d}.{int(day):02d}"


def narcotic_room_floor(room_id: str) -> str:
    if room_id in NARCOTIC_ROOM_FLOORS:
        return NARCOTIC_ROOM_FLOORS[room_id]
    if room_id.isdigit() or room_id == "RRT":
        return "5\uce35 ~ 12\uce35"
    return "\uae30\ud0c0"


def parse_narcotic_inventory(narcotic_path: Path, hospital_common_names: dict[str, str]) -> dict[str, Any]:
    wb = openpyxl.load_workbook(narcotic_path, read_only=True, data_only=True)
    ws = wb[NARCOTIC_CHECK_SHEET_NAME]
    title = clean(ws.cell(row=1, column=1).value)
    source_updated_at = narcotic_source_updated_at(title)
    header = [clean(cell.value) for cell in next(ws.iter_rows(min_row=2, max_row=2))]
    room_columns: list[tuple[int, str]] = []
    for idx, room_id in enumerate(header[3:], start=3):
        if not room_id:
            continue
        if room_id == "\ud569\uacc4":
            break
        room_columns.append((idx, room_id))

    drugs = []
    allocations = []
    categories: dict[str, str] = {}
    room_stats = {room_id: {"allocationCount": 0, "totalQuantity": 0} for _, room_id in room_columns}
    current_category = ""

    for row in ws.iter_rows(min_row=3, values_only=True):
        category = clean(row[0] if len(row) > 0 else "")
        if category in {"\ud5a5\uc815", "\ub9c8\uc57d"}:
            current_category = category
        code = clean(row[1] if len(row) > 1 else "")
        source_product_name = strip_controlled_drug_prefix(row[2] if len(row) > 2 else "")
        if not code or not source_product_name or current_category not in {"\ud5a5\uc815", "\ub9c8\uc57d"}:
            continue

        product_name = hospital_common_names.get(code, source_product_name)
        is_cold = code in NARCOTIC_COLD_STORAGE_CODES
        drugs.append(
            {
                "code": code,
                "genericName": source_product_name,
                "productName": product_name,
                "spec": "",
                "storage": "\ub0c9\uc7a5\ubcf4\uad00(2-8\u2103)" if is_cold else "\uc2e4\uc628\ubcf4\uad00",
                "note": "",
                "warning": "",
                "storageType": "REFRIGERATED" if is_cold else "ROOM",
                "narcoticCategory": current_category,
            }
        )
        categories[code] = current_category

        for idx, room_id in room_columns:
            required_qty = allocation_qty(row[idx] if idx < len(row) else None)
            if required_qty > 0:
                allocations.append({"roomId": room_id, "drugCode": code, "requiredQty": required_qty})
                room_stats[room_id]["allocationCount"] += 1
                room_stats[room_id]["totalQuantity"] += required_qty

    rooms = [
        {
            "id": room_id,
            "label": room_id,
            "sourceColumn": room_id,
            "sourceSheet": NARCOTIC_CHECK_SHEET_NAME,
            "sourceUpdatedAt": source_updated_at,
            "floor": narcotic_room_floor(room_id),
            **room_stats[room_id],
        }
        for _, room_id in room_columns
    ]
    return {
        "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "sourceFile": narcotic_path.name,
        "sourceSheet": NARCOTIC_CHECK_SHEET_NAME,
        "sourceUpdatedAt": source_updated_at,
        "drugs": drugs,
        "rooms": rooms,
        "allocations": allocations,
        "drugCategories": categories,
        "summary": {
            "drugCount": len(drugs),
            "roomCount": len(rooms),
            "allocationCount": len(allocations),
        },
    }


def normalized_name_key(name: str) -> str:
    return re.sub(r"[^0-9a-z]+", "", clean(name).lower())


def collect_warnings(stock_path: Path) -> dict[str, str]:
    wb = openpyxl.load_workbook(stock_path, read_only=True, data_only=True)
    warnings: dict[str, set[str]] = defaultdict(set)
    for ws in wb.worksheets[1:]:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [clean(v) for v in rows[0]]
        if "약품코드" not in header or "혼돈주의약품" not in header:
            continue
        code_idx = header.index("약품코드")
        warning_idx = header.index("혼돈주의약품")
        for row in rows[1:]:
            if code_idx >= len(row) or warning_idx >= len(row):
                continue
            code = clean(row[code_idx])
            warning = clean(row[warning_idx])
            if code and warning:
                warnings[code].add(warning)
    return {code: ", ".join(sorted(values)) for code, values in warnings.items()}


def parse_stock(stock_path: Path, hospital_common_names: dict[str, str]) -> dict[str, Any]:
    warnings = collect_warnings(stock_path)
    wb = openpyxl.load_workbook(stock_path, read_only=True, data_only=True)
    room_update_dates = collect_room_update_dates(wb)
    ws = wb.worksheets[0]
    header = [clean(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    room_columns = [(idx, name) for idx, name in enumerate(header[6:], start=6) if name and name != "합계"]

    drugs = []
    allocations = []
    room_stats = {name: {"allocationCount": 0, "totalQuantity": 0} for _, name in room_columns}

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_code = clean(row[0] if len(row) > 0 else "")
        source_product_name = clean(row[2] if len(row) > 2 else "")
        code = stock_code(raw_code, source_product_name)
        product_name = hospital_common_names.get(code, hospital_common_names.get(raw_code, source_product_name))
        if not raw_code or not code:
            continue
        drug = {
            "code": code,
            "genericName": clean(row[1] if len(row) > 1 else ""),
            "productName": product_name,
            "spec": clean(row[3] if len(row) > 3 else ""),
            "storage": clean(row[4] if len(row) > 4 else ""),
            "note": clean(row[5] if len(row) > 5 else ""),
            "warning": warnings.get(code, warnings.get(raw_code, clean(row[5] if len(row) > 5 else ""))),
            "storageType": storage_type(clean(row[4] if len(row) > 4 else "")),
        }
        drug.update(STOCK_FIELD_OVERRIDES.get(code, {}))
        drugs.append(drug)
        for idx, room_name in room_columns:
            required_qty = qty(row[idx] if idx < len(row) else None)
            if required_qty > 0:
                allocations.append({"roomId": room_name, "drugCode": code, "requiredQty": required_qty})
                room_stats[room_name]["allocationCount"] += 1
                room_stats[room_name]["totalQuantity"] += required_qty

    rooms = [
        {
            "id": name,
            "label": name,
            "sourceColumn": name,
            "sourceSheet": SHEET_ALIAS.get(name, name),
            "sourceUpdatedAt": room_update_dates.get(SHEET_ALIAS.get(name, name), ""),
            **room_stats[name],
        }
        for _, name in room_columns
    ]
    return {"drugs": drugs, "rooms": rooms, "allocations": allocations}


def parse_ecart(ecart_path: Path, hospital_common_names: dict[str, str]) -> dict[str, Any]:
    wb = openpyxl.load_workbook(ecart_path, read_only=True, data_only=True)
    general_ws = wb.worksheets[0]
    general_items = []
    ecart_name_aliases: dict[str, tuple[str, str]] = {}
    hospital_name_aliases: dict[str, tuple[str, str]] = {}
    for code, common_name in hospital_common_names.items():
        alias_key = normalized_name_key(common_name)
        if alias_key and alias_key not in hospital_name_aliases:
            hospital_name_aliases[alias_key] = (code, common_name)

    for row in general_ws.iter_rows(min_row=1, values_only=True):
        if not clean(row[0] if len(row) > 0 else "").isdigit():
            continue
        code = clean(row[1] if len(row) > 1 else "")
        source_name = clean(row[2] if len(row) > 2 else "")
        name = hospital_common_names.get(code, source_name)
        if not code or not name:
            continue
        item = {
            "id": code,
            "code": code,
            "name": source_name,
            "dosage": clean(row[3] if len(row) > 3 else ""),
            "quantity": qty(row[4] if len(row) > 4 else None),
        }
        item.update(ECART_FIELD_OVERRIDES.get(code, {}))
        if code in hospital_common_names:
            item["name"] = name
            for alias in {source_name, name, ECART_FIELD_OVERRIDES.get(code, {}).get("name", "")}:
                alias_key = normalized_name_key(alias)
                if alias_key:
                    ecart_name_aliases[alias_key] = (code, name)
        general_items.append(item)

    nicu_items = []
    nicu_ws = wb.worksheets[1]
    nicu_code_overrides = {
        (normalized_name_key(name), normalized_name_key(dosage)): code
        for name, dosage, code in ECART_NICU_CODE_OVERRIDES
    }
    for row in nicu_ws.iter_rows(min_row=1, values_only=True):
        no = clean(row[0] if len(row) > 0 else "")
        if not no.isdigit():
            continue
        source_name = clean(row[1] if len(row) > 1 else "")
        dosage = clean(row[2] if len(row) > 2 else "")
        if not source_name:
            continue
        alias_key = normalized_name_key(source_name)
        override_code = nicu_code_overrides.get((alias_key, normalized_name_key(dosage)))
        resolved = (
            (override_code, hospital_common_names[override_code])
            if override_code and override_code in hospital_common_names
            else ecart_name_aliases.get(alias_key) or hospital_name_aliases.get(alias_key)
        )
        code = resolved[0] if resolved else ""
        name = resolved[1] if resolved else source_name
        nicu_items.append(
            {
                "id": f"NICU-{int(no):02d}",
                "code": code,
                "name": name,
                "dosage": dosage,
                "quantity": qty(row[4] if len(row) > 4 else None),
            }
        )

    departments = []
    receive_ws = wb.worksheets[2]
    for row in receive_ws.iter_rows(min_row=3, values_only=True):
        value = clean(row[0] if len(row) > 0 else "")
        if value:
            departments.append(value)
    return {"generalItems": general_items, "nicuItems": nicu_items, "departments": departments}


def parse_checklist(checklist_path: Path) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(checklist_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    items = []
    current_section = ""
    for row in ws.iter_rows(values_only=True):
        cells = [clean(value) for value in row]
        nonempty = [cell for cell in cells if cell and cell != "□"]
        if not nonempty:
            continue
        first = nonempty[0]
        if first.startswith("[") and first.endswith("]"):
            current_section = first.strip("[]")
            rest = " ".join(nonempty[1:])
            if rest and not is_checklist_label_row(rest) and not is_retired_checklist_row(rest):
                items.append({"section": current_section, "text": rest})
            continue
        if "병동 비품약&E-cart 점검 체크리스트" in first or "점검 내용을" in first:
            continue
        text = " ".join(nonempty)
        if text.startswith("4. 비품이외의 잉여약을 보관하고 있다."):
            text = "4. 비품이외의 잉여약을 보관하고 있지 않다."
        if text and current_section and not is_checklist_label_row(text) and not is_retired_checklist_row(text):
            if text.startswith("2-1 ") and " 2-2 " in text:
                first, second = text.split(" 2-2 ", 1)
                items.append({"section": current_section, "text": first})
                items.append({"section": current_section, "text": f"2-2 {second}"})
            else:
                items.append({"section": current_section, "text": text})
    if not any(item["section"] == "냉장약" and item["text"].startswith("6.") for item in items):
        insert_at = max((idx for idx, item in enumerate(items) if item["section"] == "냉장약"), default=-1) + 1
        items.insert(insert_at, {"section": "냉장약", "text": "6. 연 1회 냉장고 온도계 검증 여부"})
    return items


def main() -> None:
    stock_path = find_workbook("202606")
    ecart_path = find_workbook("E-cart")
    checklist_path = find_workbook("체크리스트")
    hospital_path = HOSPITAL_DRUG_LIST_PATH
    if not hospital_path.exists():
        raise FileNotFoundError(f"Cannot find hospital drug workbook: {hospital_path}")
    narcotic_path = find_workbook(NARCOTIC_STATUS_KEYWORD)
    hospital_common_names = load_hospital_common_names(hospital_path)
    stock = parse_stock(stock_path, hospital_common_names)
    ecart = parse_ecart(ecart_path, hospital_common_names)
    checklist = parse_checklist(checklist_path)
    narcotic = parse_narcotic_inventory(narcotic_path, hospital_common_names)
    data = {
        "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "sourceFiles": {
            "stockWorkbook": stock_path.name,
            "ecartWorkbook": ecart_path.name,
            "checklistWorkbook": checklist_path.name,
        },
        "summary": {
            "stockDrugCount": len(stock["drugs"]),
            "stockRoomCount": len(stock["rooms"]),
            "stockAllocationCount": len(stock["allocations"]),
            "ecartGeneralItemCount": len(ecart["generalItems"]),
            "ecartNicuItemCount": len(ecart["nicuItems"]),
            "ecartDepartmentCount": len(ecart["departments"]),
            "checklistItemCount": len(checklist),
        },
        "stock": stock,
        "ecart": ecart,
        "checklist": checklist,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    NARCOTIC_INVENTORY_OUT.write_text(json.dumps(narcotic, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(data["summary"], ensure_ascii=False, indent=2))
    print(f"wrote {OUT.relative_to(ROOT)}")
    print(json.dumps(narcotic["summary"], ensure_ascii=False, indent=2))
    print(f"wrote {NARCOTIC_INVENTORY_OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
