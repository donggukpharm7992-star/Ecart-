from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "원내보유의약품_라벨매칭_20260702.xlsx"
OUTPUT = ROOT / "src" / "data" / "pharmacyLabelMatches.generated.json"

TEXT_FIELDS = {
    "약품코드": "code",
    "상용약품명": "englishName",
    "한글약품명": "koreanName",
    "함량": "strength",
    "규격": "spec",
    "포장": "package",
    "보관법": "storage",
    "해당라벨": "matchedLabel",
    "라벨원본파일": "sourceFile",
    "라벨시트/문서위치": "sourceLocation",
    "매칭상태": "matchStatus",
    "매칭근거": "matchReason",
    "조건출처": "conditionSource",
    "확인메모": "reviewMemo",
}

BOOL_FIELDS = {
    "차광": "lightProtected",
    "냉장": "refrigerated",
    "용량주의": "doseCaution",
    "유사발음": "similarSound",
    "유사모양": "similarLook",
    "고위험의약품": "highRisk",
    "고주의약품": "highCaution",
    "항암제": "anticancer",
    "마약": "narcotic",
    "향정": "psychotropic",
    "고가통계약": "highCost",
    "이름주의": "nameCaution",
}


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("_x000D_", "").strip()


def is_checked(value: object) -> bool:
    text = clean(value).upper()
    return text in {"Y", "YES", "TRUE", "1", "O", "○", "예", "해당", "차광", "냉장"}


def score(value: object) -> int:
    text = clean(value)
    if not text:
        return 0
    return int(float(text))


def main() -> None:
    if not SOURCE.exists():
        raise FileNotFoundError(f"Missing source workbook: {SOURCE}")

    workbook = load_workbook(SOURCE, data_only=True, read_only=True)
    worksheet = workbook["라벨매칭"]
    headers = [clean(value).replace("\n", " ") for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {header: position for position, header in enumerate(headers)}

    rows: list[dict[str, object]] = []
    for raw in worksheet.iter_rows(min_row=2, values_only=True):
        code = clean(raw[index["약품코드"]])
        if not code:
            continue

        row: dict[str, object] = {}
        for header, key in TEXT_FIELDS.items():
            row[key] = clean(raw[index[header]]) if header in index else ""
        for header, key in BOOL_FIELDS.items():
            row[key] = is_checked(raw[index[header]]) if header in index else False
        row["matchScore"] = score(raw[index["매칭점수"]]) if "매칭점수" in index else 0
        rows.append(row)

    rows.sort(key=lambda item: (str(item["englishName"]).lower(), str(item["code"]).lower()))
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(rows, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(rows)} pharmacy label match rows to {OUTPUT}")


if __name__ == "__main__":
    main()
