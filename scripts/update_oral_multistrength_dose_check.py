from __future__ import annotations

import re
import shutil
import tempfile
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "약제팀 라벨" / "원내보유의약품리스트.xlsx"
BACKUP = Path("H:/CHOI/라벨앱/원내보유의약품리스트_백업.xlsx")


def clean(value: object) -> str:
    return str(value or "").replace("_x000D_", "").strip()


def core_name(value: object) -> str:
    text = clean(value).lower()
    text = re.sub(
        r"\d+(?:\.\d+)?\s*(?:mcg|μg|mg|g|ml|iu|%)(?:/\d+(?:\.\d+)?\s*(?:ml|g))?",
        "",
        text,
    )
    text = re.sub(r"\b(?:tab(?:let)?|cap(?:sule)?)\b", "", text)
    return re.sub(r"[^0-9a-z가-힣]", "", text)


def main() -> None:
    workbook = load_workbook(SOURCE)
    worksheet = workbook.worksheets[0]
    headers = [clean(cell.value).replace("\n", " ") for cell in worksheet[1]]
    index = {header: position + 1 for position, header in enumerate(headers)}
    groups: dict[str, list[int]] = defaultdict(list)

    for row_number in range(2, worksheet.max_row + 1):
        drug_type = clean(worksheet.cell(row_number, index["약품유형"]).value).replace(" ", "")
        in_hospital = clean(worksheet.cell(row_number, index["원내보유"]).value).upper()
        name = worksheet.cell(row_number, index["상용약품명"]).value
        if drug_type not in {"원병", "PTP"} or in_hospital != "Y" or not clean(name):
            continue
        groups[core_name(name)].append(row_number)

    targets: list[int] = []
    for row_numbers in groups.values():
        if len(row_numbers) < 2:
            continue
        for row_number in row_numbers:
            dose_caution = clean(worksheet.cell(row_number, index["용량주의"]).value).upper()
            dose_check = clean(worksheet.cell(row_number, index["용량확인"]).value).upper()
            if dose_caution != "Y" and dose_check != "Y":
                worksheet.cell(row_number, index["용량확인"]).value = "Y"
                targets.append(row_number)

    BACKUP.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SOURCE, BACKUP)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", dir=SOURCE.parent, delete=False) as temporary:
        temporary_path = Path(temporary.name)
    try:
        workbook.save(temporary_path)
        verified = load_workbook(temporary_path, read_only=True, data_only=False)
        if len(verified.sheetnames) != len(workbook.sheetnames):
            raise RuntimeError("저장 검증 중 시트 수가 달라졌습니다.")
        verified.close()
        temporary_path.replace(SOURCE)
    finally:
        temporary_path.unlink(missing_ok=True)

    print(f"Updated {len(targets)} dose-check cells across {sum(len(rows) > 1 for rows in groups.values())} multi-strength groups.")


if __name__ == "__main__":
    main()
