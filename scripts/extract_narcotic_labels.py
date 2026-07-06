from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "마약"
OUTPUT = ROOT / "src" / "data" / "narcoticLabels.generated.json"

LABEL_WORKBOOKS = [
    ("향정라벨.xlsx", "향정"),
    ("마약주사라벨.xlsx", "마약"),
    ("마약경구라벨.xlsx", "마약"),
]

LABEL_COLUMN_PAIRS = [(1, 2), (3, 4)]


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("_x000D_", "").strip()


def clean_multiline(value: object) -> str:
    return "\n".join(part.strip() for part in clean(value).splitlines() if part.strip())


def main() -> None:
    rows: list[dict[str, str]] = []
    seen: set[tuple[str, str, str]] = set()

    for file_name, category in LABEL_WORKBOOKS:
        source = SOURCE_DIR / file_name
        if not source.exists():
            raise FileNotFoundError(f"Missing narcotic label workbook: {source}")

        workbook = load_workbook(source, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]

        for row_number in range(1, worksheet.max_row + 1):
            for code_column, label_column in LABEL_COLUMN_PAIRS:
                code = clean(worksheet.cell(row=row_number, column=code_column).value)
                label_text = clean_multiline(worksheet.cell(row=row_number, column=label_column).value)
                if not code or code == "약품코드" or not label_text:
                    continue

                key = (code, label_text, file_name)
                if key in seen:
                    continue
                seen.add(key)

                rows.append(
                    {
                        "code": code,
                        "labelText": label_text,
                        "category": category,
                        "categoryText": clean_multiline(worksheet.cell(row=row_number + 1, column=label_column).value),
                        "cautionText": clean_multiline(worksheet.cell(row=row_number + 2, column=label_column).value),
                        "sourceFile": file_name,
                        "sourceSheet": worksheet.title,
                        "sourceCell": f"{get_column_letter(label_column)}{row_number}",
                    }
                )

        workbook.close()

    rows.sort(key=lambda row: (row["category"], row["labelText"].lower(), row["code"].lower(), row["sourceFile"]))
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(rows, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(rows)} narcotic label rows to {OUTPUT}")


if __name__ == "__main__":
    main()
