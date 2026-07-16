from __future__ import annotations

import shutil
import tempfile
import zipfile
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "약제팀 라벨" / "원내보유의약품리스트.xlsx"
BACKUP_DIR = Path("H:/CHOI/라벨앱")
BACKUP = BACKUP_DIR / "원내보유의약품리스트_백업.xlsx"
SHEET_XML = "xl/worksheets/sheet1.xml"


def validate(path: Path) -> tuple[int, int]:
    with zipfile.ZipFile(path) as archive:
        bad_entry = archive.testzip()
        if bad_entry:
            raise RuntimeError(f"XLSX 압축 항목 오류: {bad_entry}")
        sheet_xml = archive.read(SHEET_XML)
        if b'Ignorable="x14ac"' in sheet_xml and b"xmlns:x14ac=" not in sheet_xml:
            raise RuntimeError("약품조회 시트의 x14ac XML 선언이 누락되었습니다.")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if workbook.sheetnames != ["약품조회", "라벨 생성규칙", "경구 주사 리스트", "영양수액리스트", "외용제리스트", "시럽리스트"]:
            raise RuntimeError(f"예상하지 못한 시트 구성입니다: {workbook.sheetnames}")
        worksheet = workbook["약품조회"]
        if worksheet.max_row < 2 or worksheet.max_column < 41:
            raise RuntimeError(f"약품조회 범위가 비정상입니다: {worksheet.max_row}행 {worksheet.max_column}열")
        return worksheet.max_row - 1, worksheet.max_column
    finally:
        workbook.close()


def main() -> None:
    drug_count, column_count = validate(SOURCE)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=BACKUP_DIR) as temporary:
        temporary_path = Path(temporary.name)
    try:
        shutil.copy2(SOURCE, temporary_path)
        validate(temporary_path)
        temporary_path.replace(BACKUP)
    finally:
        temporary_path.unlink(missing_ok=True)
    print(f"Backup updated: {BACKUP} ({drug_count} drugs, {column_count} columns)")


if __name__ == "__main__":
    main()
