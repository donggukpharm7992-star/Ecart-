from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(__file__).resolve().parent.parent / "약제팀 라벨" / "원내보유의약품리스트.xlsx"
HEADER = "일반수액 색기호"
FLUID_COLORS = {
    "XAQD": "green",
    "XD10W": "black",
    "XD10W5": "black",
    "XD15W": "black",
    "XD20W1L": "black",
    "XD50W1": "black",
    "XD5S": "pink",
    "XD5S5": "pink",
    "XD5W": "black",
    "XD5W100": "black",
    "XD5W2Y": "black",
    "XD5W5": "black",
    "XD5W50": "orange",
    "XDNK1": "purple",
    "XDNK2": "red",
    "XDNK25": "red",
    "XDNK35": "black",
    "XGD20W3": "black",
    "XGNS150": "blue",
    "XHD": "yellow",
    "XHNS": "green",
    "XHS": "orange",
    "XHS5": "orange",
    "XNAK40": "orange",
    "XNS100": "black",
    "XNS1L": "blue",
    "XNS250": "pink",
    "XNS3L": "blue",
    "XNS500": "blue",
    "XNS50C": "green",
    "XPLASMA": "green",
    "XPLASMA5": "green",
}


def copy_cell_style(source, target) -> None:
    target._style = copy(source._style)
    target.number_format = source.number_format
    target.protection = copy(source.protection)
    target.alignment = copy(source.alignment)
    target.fill = copy(source.fill)
    target.font = copy(source.font)
    target.border = copy(source.border)


def main() -> None:
    workbook = load_workbook(WORKBOOK)
    worksheet = workbook["약품조회"]
    headers = {cell.value: cell.column for cell in worksheet[1]}
    if "약품코드" not in headers or "약품유형" not in headers:
        raise ValueError("약품조회 시트에 약품코드 또는 약품유형 열이 없습니다.")

    color_column = headers.get(HEADER)
    if color_column is None:
        color_column = worksheet.max_column + 1
        header_cell = worksheet.cell(1, color_column)
        source_header = worksheet.cell(1, headers["약품유형"])
        copy_cell_style(source_header, header_cell)
        header_cell.value = HEADER
        worksheet.column_dimensions[header_cell.column_letter].width = 16

    code_column = headers["약품코드"]
    type_column = headers["약품유형"]
    updated = 0
    fluid_codes: set[str] = set()
    for row_number in range(2, worksheet.max_row + 1):
        code = str(worksheet.cell(row_number, code_column).value or "").strip().upper()
        drug_type = str(worksheet.cell(row_number, type_column).value or "").strip()
        color_cell = worksheet.cell(row_number, color_column)
        if drug_type != "일반수액":
            continue
        if code not in FLUID_COLORS:
            raise ValueError(f"일반수액 색기호가 정의되지 않은 약품코드: {code}")
        if color_cell.value != FLUID_COLORS[code]:
            color_cell.value = FLUID_COLORS[code]
            updated += 1
        fluid_codes.add(code)

    if fluid_codes != set(FLUID_COLORS):
        missing = sorted(set(FLUID_COLORS) - fluid_codes)
        unexpected = sorted(fluid_codes - set(FLUID_COLORS))
        raise ValueError(f"일반수액 코드 대조 실패: 누락={missing}, 추가={unexpected}")

    temporary = WORKBOOK.with_suffix(".tmp.xlsx")
    workbook.save(temporary)
    temporary.replace(WORKBOOK)
    print(f"Updated {updated} fluid color symbols in {WORKBOOK.name}.")


if __name__ == "__main__":
    main()
