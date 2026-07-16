from __future__ import annotations

import json
import hashlib
import re
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook


LABEL_DIR = Path(__file__).resolve().parent
SOURCE = LABEL_DIR / "원내보유의약품리스트.xlsx"
OUTPUT = LABEL_DIR / "data" / "hospitalDrugLabels.generated.json"
SIDE_LABEL_TEMPLATE = LABEL_DIR / "뺑뺑이 PTP통, 병_측면라벨_양식.xlsx"
IMAGE_OUTPUT_DIR = LABEL_DIR.parent / "public" / "pharmacy-drug-images"


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("_x000D_", "").strip()


def is_yes(value: object) -> bool:
    return clean(value).upper() == "Y"


def normalized_name(value: object) -> str:
    return re.sub(r"[^0-9a-z가-힣]", "", clean(value).lower())


def extract_side_label_images() -> dict[str, str]:
    if not SIDE_LABEL_TEMPLATE.exists():
        return {}
    workbook = load_workbook(SIDE_LABEL_TEMPLATE, data_only=True)
    IMAGE_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    image_by_name: dict[str, str] = {}
    for worksheet in workbook.worksheets:
        for image in worksheet._images:
            anchor_row = image.anchor._from.row + 1
            name_column = image.anchor._from.col + 2
            label_text = ""
            for row_number in range(max(1, anchor_row - 2), min(worksheet.max_row, anchor_row + 3) + 1):
                candidate = clean(worksheet.cell(row_number, name_column).value)
                if candidate:
                    label_text = candidate
                    break
            if not label_text:
                continue
            names = [part.strip("() ") for part in label_text.splitlines()[:2] if part.strip()]
            if not names:
                continue
            data = image._data()
            extension = image.format.lower() if image.format else "png"
            digest = hashlib.sha1(data).hexdigest()[:16]
            file_name = f"{digest}.{extension}"
            target = IMAGE_OUTPUT_DIR / file_name
            if not target.exists():
                target.write_bytes(data)
            public_path = f"/pharmacy-drug-images/{file_name}"
            for name in names:
                image_by_name[normalized_name(name)] = public_path
    return image_by_name


def match_image(image_by_name: dict[str, str], *names: str) -> str:
    normalized = [normalized_name(name) for name in names if clean(name)]
    for name in normalized:
        if name in image_by_name:
            return image_by_name[name]
    candidates = [
        (len(image_name), image_path)
        for image_name, image_path in image_by_name.items()
        if len(image_name) >= 4 and any(image_name in name or name in image_name for name in normalized)
    ]
    return max(candidates, default=(0, ""))[1]


def main() -> None:
    image_by_name = extract_side_label_images()
    workbook = load_workbook(SOURCE, data_only=True, read_only=True)
    worksheet = workbook.worksheets[0]
    oral_injection_names = {
        clean(value).lower()
        for row in workbook.worksheets[2].iter_rows(min_row=2, values_only=True)
        for value in (row[1], row[5], row[9])
        if clean(value)
    }
    nutrition_names = {
        clean(row[0]).lower() for row in workbook.worksheets[3].iter_rows(min_row=2, values_only=True) if clean(row[0])
    }
    external_codes = {
        clean(row[0]) for row in workbook.worksheets[4].iter_rows(min_row=2, values_only=True) if clean(row[0])
    }
    syrup_codes = {
        clean(row[0]) for row in workbook.worksheets[5].iter_rows(min_row=2, values_only=True) if clean(row[0])
    }
    headers = [clean(value).replace("\n", " ") for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {header: position for position, header in enumerate(headers)}

    def read(raw: tuple[object, ...], header: str) -> str:
        return clean(raw[index[header]])

    def read_optional(raw: tuple[object, ...], header: str) -> str:
        return clean(raw[index[header]]) if header in index else ""

    rows: list[dict[str, object]] = []
    for raw in worksheet.iter_rows(min_row=2, values_only=True):
        code = read(raw, "약품코드")
        name = read(raw, "상용약품명")
        if not code or not name:
            continue
        rows.append(
            {
                "code": code,
                "itemCode": read(raw, "물품코드"),
                "name": name,
                "koreanName": read(raw, "한글약품명"),
                "strength": read(raw, "함량"),
                "drugType": read(raw, "약품유형"),
                "highCost": is_yes(raw[index["고가약"]]),
                "spec": read(raw, "규격"),
                "package": read(raw, "포장"),
                "storage": read(raw, "보관법"),
                "lightProtected": read(raw, "차광필요") == "차광",
                "inHospital": is_yes(raw[index["원내보유"]]),
                "oralAnticancer": is_yes(raw[index["경구항암제"]]),
                "similarLook": is_yes(raw[index["유사모양"]]),
                "similarSound": is_yes(raw[index["유사발음"]]),
                "doseCaution": is_yes(raw[index["용량주의"]]),
                "doseCheck": is_yes(raw[index["용량확인"]]),
                "highRisk": is_yes(raw[index["고위험의약품"]]),
                "highRiskCategory": read(raw, "고위험의약품분류"),
                "atc": read(raw, "ATC"),
                "ptpOpened": is_yes(raw[index["PTP깐거"]]),
                "inpatientPowderPtp": is_yes(raw[index["입원산제용PTP 깐거"]]),
                "threeTierHalf": is_yes(raw[index["3단반알"]]),
                "expiry": read(raw, "유효기간"),
                "location": read(raw, "위치"),
                "ampouleHolder": read(raw, "앰플꽂이"),
                "sideLabel1T": read(raw, "1T 3단장 뺑뺑이 PTP 측면라벨"),
                "sideLabelHalfT": read(raw, "0.5T 3단장 뺑뺑이 병 측면라벨"),
                "sideLabelQuarterT": read(raw, "0.25T 3단장 뺑뺑이 병 측면라벨"),
                "coloredSideLabel": read(raw, "3단장 유색 반티통 측면라벨"),
                "coloredSideBackground": read(raw, "3단장 유색 반티통 측면라벨 바탕색"),
                "capLabel": read(raw, "3단장 유색 반티통 병뚜껑"),
                "capBackground": read(raw, "3단장 반티통 병뚜껑 바탕색 기호"),
                "nameCaution": is_yes(raw[index["이름주의"]]),
                "border": is_yes(raw[index["테두리"]]),
                "borderColor": read(raw, "테두리 색기호"),
                "cabinetOralInjection": name.lower() in oral_injection_names,
                "cabinetNutrition": name.lower() in nutrition_names,
                "cabinetExternal": code in external_codes,
                "cabinetSyrup": code in syrup_codes,
                "imagePath": read_optional(raw, "식별사진경로") or match_image(image_by_name, name, read(raw, "한글약품명")),
                "imageSourceUrl": read_optional(raw, "식별사진출처") or f"https://www.health.kr/searchDrug/search_detail.asp?search_detail=Y&search_sunb1={quote(read(raw, '한글약품명'))}",
            }
        )

    rows.sort(key=lambda row: (str(row["name"]).lower(), str(row["code"]).lower()))
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(rows, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(rows)} hospital drug label rows to {OUTPUT}")


if __name__ == "__main__":
    main()
