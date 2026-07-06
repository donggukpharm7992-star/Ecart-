from __future__ import annotations
import sys
import io
from pathlib import Path
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

FILE = Path(__file__).resolve().parent / "비치향정,마약현황.xlsx"

def main() -> None:
    wb = load_workbook(FILE, data_only=True)
    
    # Just print the first sheet (점검) fully
    ws = wb[wb.sheetnames[0]]
    sheet_name = wb.sheetnames[0]
    print(f"Sheet: '{sheet_name}'")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    
    if ws.merged_cells.ranges:
        print(f"\nMerged cell ranges ({len(ws.merged_cells.ranges)}):")
        for mc in ws.merged_cells.ranges:
            print(f"  {mc}")
    
    print(f"\n--- All data ---")
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False), start=1):
        values = []
        for cell in row:
            values.append(cell.value)
        print(f"Row {row_idx:3d}: {values}")
    
    wb.close()

if __name__ == "__main__":
    main()
